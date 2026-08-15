"""表格工具（本地 Excel/CSV：读/写/预览，支持按表头定位）。"""
import csv
from pathlib import Path

from config.settings import settings
from backend.tools.base import BaseTool, ToolResult
from backend.tools.registry import register

DATA_DIR = Path(settings.sheets_data_dir)


@register
class SheetTool(BaseTool):
    """本地表格工具：只允许操作 DATA_DIR 内文件。"""

    name = "sheets"
    high_risk = False
    description = '本地表格：读取/按表头定位写入（高危）/预览（高危）'
    args_schema = {
        "type": "object",
        "required": ["action"],
        "properties": {"action": {"type": "string", "enum": ["read", "write", "write_by_key", "preview"]}},
        "oneOf": [
            {"properties": {"action": {"const": "read"}, "filename": {"type": "string"}, "limit": {"type": "integer"}}, "required": ["action", "filename"]},
            {"properties": {"action": {"const": "write_by_key"}, "filename": {"type": "string"}, "key_column": {"type": "string"}, "key_value": {"type": "string"}, "field": {"type": "string"}, "value": {"type": "string"}}, "required": ["action", "filename", "key_column", "key_value", "field", "value"]},
            {"properties": {"action": {"const": "write"}, "filename": {"type": "string"}, "changes": {"type": "array"}}, "required": ["action", "filename", "changes"]},
            {"properties": {"action": {"const": "preview"}, "filename": {"type": "string"}, "changes": {"type": "array"}}, "required": ["action", "filename", "changes"]},
        ],
    }  # 写操作由高危关键词判定（更新/写入=高危）

    def execute(self, action: str = "", **kwargs) -> ToolResult:
        try:
            if action == "read":
                return self.read(filename=kwargs.get("filename", ""), limit=int(kwargs.get("limit", 10)))
            if action == "preview":
                return self.preview(filename=kwargs.get("filename", ""), changes=kwargs.get("changes", []))
            if action == "write":
                return self.write(filename=kwargs.get("filename", ""), changes=kwargs.get("changes", []))
            if action == "write_by_key":
                return self.write_by_key(
                    filename=kwargs.get("filename", ""),
                    key_column=kwargs.get("key_column", "姓名"),
                    key_value=kwargs.get("key_value", ""),
                    field=kwargs.get("field", ""),
                    value=kwargs.get("value", ""),
                )
            return ToolResult(ok=False, message=f"不支持的表格动作：{action}")
        except ValueError as exc:
            return ToolResult(ok=False, message=str(exc))

    def _resolve(self, filename: str) -> Path:
        if not filename:
            raise ValueError("缺少文件名")
        base = DATA_DIR.resolve()
        path = (base / filename).resolve()
        if not path.is_relative_to(base):
            raise ValueError(f"路径越权：只允许操作 {base} 内文件")
        return path

    def read(self, filename: str, limit: int = 10) -> ToolResult:
        path = self._resolve(filename)
        if not path.exists():
            return ToolResult(ok=False, message=f"文件不存在：{filename}")
        rows = _read_rows(path, limit)
        return ToolResult(ok=True, message=f"读取 {len(rows)} 行", data={"file": filename, "rows": rows})

    def find_cell(self, filename: str, key_column: str, key_value: str, field: str, header_row: int = 1) -> dict:
        """按表头定位单元格：在 key_column 列找到 key_value 的行，取 field 列。返回 {row, column, old}。"""
        path = self._resolve(filename)
        if not path.exists():
            raise ValueError(f"文件不存在：{filename}")
        rows = _read_rows(path, 500)
        if not rows:
            raise ValueError("表格为空")
        header = [str(h) if h is not None else "" for h in rows[header_row - 1]]
        key_col = _header_index(header, key_column)
        field_col = _header_index(header, field)
        if key_col is None:
            raise ValueError(f"找不到表头：{key_column}")
        if field_col is None:
            raise ValueError(f"找不到表头：{field}")
        for i, row in enumerate(rows[header_row:], start=header_row + 1):
            cell = row[key_col - 1] if len(row) >= key_col else None
            if cell is not None and str(cell).strip() == str(key_value).strip():
                old = row[field_col - 1] if len(row) >= field_col else ""
                return {"row": i, "column": _col_letter(field_col), "old": "" if old is None else str(old)}
        raise ValueError(f"找不到 {key_column}={key_value}")

    def preview(self, filename: str, changes: list) -> ToolResult:
        path = self._resolve(filename)
        if not path.exists():
            return ToolResult(ok=False, message=f"文件不存在：{filename}")
        rows = _read_rows(path, 200)
        preview = []
        for c in changes:
            old = _cell_value(rows, c.get("row"), c.get("column"))
            preview.append({"row": c.get("row"), "column": c.get("column"), "old": old, "new": c.get("value", "")})
        return ToolResult(ok=True, message=f"生成 {len(preview)} 处变更预览", data={"preview": preview})

    def write(self, filename: str, changes: list) -> ToolResult:
        path = self._resolve(filename)
        if not path.exists():
            return ToolResult(ok=False, message=f"文件不存在：{filename}")
        count = _apply_changes(path, changes)
        return ToolResult(ok=True, message=f"已更新 {count} 个单元格", data={"file": filename})

    def write_by_key(self, filename: str, key_column: str, key_value: str, field: str, value: str) -> ToolResult:
        """按表头+关键值定位后写入（先定位，找不到就报错，不瞎改）。"""
        target = self.find_cell(filename, key_column, key_value, field)
        return self.write(filename, [{"row": target["row"], "column": target["column"], "value": value}])


def _read_rows(path: Path, limit: int) -> list:
    if path.suffix.lower() == ".csv":
        with open(path, "r", encoding="utf-8", newline="") as f:
            return [row for row in csv.reader(f)][:limit]
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    return [[ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)] for r in range(1, ws.max_row + 1)][:limit]


def _header_index(header: list[str], name: str) -> int | None:
    for idx, h in enumerate(header, start=1):
        if h.strip() == name.strip():
            return idx
    return None


def _col_letter(idx: int) -> str:
    result = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        result = chr(ord("A") + rem) + result
    return result


def _cell_value(rows: list, row, column) -> str:
    try:
        idx = _col_index(column) - 1
        value = rows[row - 1][idx]
        return "" if value is None else str(value)
    except (IndexError, TypeError):
        return ""


def _apply_changes(path: Path, changes: list) -> int:
    if path.suffix.lower() == ".csv":
        with open(path, "r", encoding="utf-8", newline="") as f:
            rows = list(csv.reader(f))
        for c in changes:
            row, col, value = int(c["row"]), _col_index(c["column"]), str(c.get("value", ""))
            while len(rows) < row:
                rows.append([])
            while len(rows[row - 1]) < col:
                rows[row - 1].append("")
            rows[row - 1][col - 1] = value
        with open(path, "w", encoding="utf-8", newline="") as f:
            csv.writer(f).writerows(rows)
        return len(changes)

    import openpyxl

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    for c in changes:
        ws.cell(row=int(c["row"]), column=_col_index(c["column"]), value=str(c.get("value", "")))
    wb.save(path)
    return len(changes)


def _col_index(column) -> int:
    if isinstance(column, int):
        return column
    col = str(column).upper()
    result = 0
    for ch in col:
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result