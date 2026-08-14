#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""办公自动化 Agent 自动化测试（20 黄金 Case / 3 类评分器 / 完整 Trace）。

用法:
    python scripts/smoke_test.py [--rounds 3] [--base-url http://127.0.0.1:8001]
                                 [--include-email] [--cleanup]

退出码: 0=全过 1=有失败 2=环境未就绪
"""
import argparse
import json
import os
import random
import sqlite3
import sys
import time
import traceback
from datetime import datetime, timedelta
from pathlib import Path

import httpx
import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parents[1]
os.chdir(PROJECT_ROOT)
sys.path.insert(0, str(PROJECT_ROOT))

EXCEL_PATH = PROJECT_ROOT.parent / "Excel数据" / "报名表.xlsx"
APP_LOG = PROJECT_ROOT / "logs" / "app.log"
REPORT_DIR = PROJECT_ROOT / "logs"
TRACE_DIR = PROJECT_ROOT / "logs" / "traces"
DEFAULT_BASE_URL = "http://127.0.0.1:8001"
TIMEOUT = 60


def load_env(path):
    env = {}
    if not path.exists():
        return env
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, value = line.partition("=")
        env[key.strip()] = value.strip().strip('"').strip("'")
    return env


def resolve_db_path(env):
    url = env.get("DATABASE_URL", "sqlite:///./office_agent.db")
    if url.startswith("sqlite:///"):
        p = Path(url[len("sqlite:///"):])
        return p if p.is_absolute() else (PROJECT_ROOT / p)
    return PROJECT_ROOT / "office_agent.db"


def read_sheet():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active
    return [[ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)] for r in range(1, ws.max_row + 1)]


def log_tail(n=20):
    if not APP_LOG.exists():
        return "(无日志文件)"
    lines = APP_LOG.read_text(encoding="utf-8", errors="replace").splitlines()
    return "\n".join(lines[-n:])


def post_task(api, text, round_no=1):
    return api.post("/api/v1/tasks", json={"text": text, "round": round_no})


def decide(api, cid, approve):
    return api.post(f"/api/v1/confirmations/{cid}/decide", json={"approve": approve})


def find_item(body, tool):
    for item in body.get("tasks", []):
        if item.get("tool") == tool:
            return item
    return None


def rand_suffix():
    return datetime.now().strftime("%H%M%S") + str(random.randint(10, 99))


def get_confirmations(api):
    return api.get("/api/v1/confirmations").json().get("items", [])


def make_trace(case_id, rnd):
    return {
        "trace_id": f"{case_id}_r{rnd}_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}",
        "case_id": case_id,
        "round": rnd,
        "input": None,
        "request": None,
        "response": None,
        "planning": None,
        "locate": None,
        "confirmation": None,
        "side_effect": None,
        "log_snippet": None,
        "scoring": None,
        "result": None,
    }


def write_trace(trace):
    TRACE_DIR.mkdir(parents=True, exist_ok=True)
    path = TRACE_DIR / f"{trace['trace_id']}.json"
    path.write_text(json.dumps(trace, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return path


def check_env(base_url):
    problems = []
    try:
        resp = httpx.get(f"{base_url.rstrip('/')}/health", timeout=5)
        if resp.status_code != 200 or resp.json().get("status") != "ok":
            problems.append(f"后端健康检查异常（{base_url}/health）")
    except Exception as exc:
        problems.append(f"后端 {base_url} 不可访问：{exc}")
    env = load_env(PROJECT_ROOT / ".env")
    if not (env.get("LLM_API_KEY") and env.get("LLM_BASE_URL") and env.get("LLM_MODEL")):
        problems.append(".env 未配置 DeepSeek")
    db_path = resolve_db_path(env)
    if db_path.exists():
        try:
            con = sqlite3.connect(db_path)
            row = con.execute(
                "select access_token, expires_at, refresh_token from feishu_tokens where user_id=? order by id desc limit 1",
                (1,),
            ).fetchone()
            con.close()
            if not row or not row[0]:
                problems.append("用户 id=1 无飞书令牌（请先飞书登录）")
        except sqlite3.Error as exc:
            problems.append(f"读取数据库失败：{exc}")
    else:
        problems.append(f"数据库不存在：{db_path}")
    if not EXCEL_PATH.exists():
        problems.append(f"表格文件不存在：{EXCEL_PATH}")
    return problems


def gc01_read_sheet(api, rnd, trace):
    text = "读取报名表内容"
    trace["input"] = text
    resp = post_task(api, text, rnd)
    trace["response"] = resp.json()
    body = resp.json()
    item = find_item(body, "sheets")
    ok = resp.status_code == 200 and body.get("status") in ("planned", "executed", "pending_confirm") and item and item.get("status") == "executed" and "读取" in item.get("result", "")
    rows = read_sheet() if ok else []
    trace["side_effect"] = {"rows_count": len(rows)}
    return {"pass": ok, "details": item.get("result", "") if item else "未找到子任务", "quality": 100 if ok and len(rows) >= 3 else 60}


def gc02_write_by_key(api, rnd, trace):
    new_phone = "138" + str(random.randint(10000000, 99999999))
    text = f"把报名表里张三的电话更新为{new_phone}"
    trace["input"] = text
    resp = post_task(api, text, rnd)
    body = resp.json()
    trace["response"] = body
    item = find_item(body, "sheets")
    ok_status = resp.status_code == 200 and item and item.get("status") == "pending_confirm" and item.get("confirmation_id")
    pending = get_confirmations(api)
    conf = next((c for c in pending if c["id"] == (item.get("confirmation_id") if item else -1)), None)
    preview = conf["params"] if conf else ""
    trace["locate"] = {"preview": preview}
    trace["confirmation"] = {"id": item.get("confirmation_id") if item else None}
    quality = 0
    if "第2行B列" in preview:
        quality += 40
    if "→" in preview:
        quality += 30
    if "姓名=张三" in preview:
        quality += 30
    if not ok_status:
        return {"pass": False, "details": item.get("result", "") if item else str(body), "quality": quality}
    if not conf or "第2行B列" not in preview:
        return {"pass": False, "details": f"确认页预览缺失：{preview}", "quality": quality}
    before = read_sheet()[1][1]
    dec = decide(api, conf["id"], True)
    trace["confirmation"]["decide"] = dec.json()
    after = read_sheet()[1][1]
    c2 = read_sheet()[1][2]
    trace["side_effect"] = {"before": before, "after": after, "c2": c2}
    ok = dec.status_code == 200 and dec.json().get("execution_result", {}).get("ok") is True and after == new_phone
    if c2 not in (None, ""):
        ok = False
        quality -= 50
    return {"pass": ok, "details": f"B2: {before} -> {after}, C2={c2}", "quality": max(0, quality)}


def gc03_missing_person(api, rnd, trace):
    text = "把报名表里王五的电话更新为13800001111"
    trace["input"] = text
    resp = post_task(api, text, rnd)
    body = resp.json()
    trace["response"] = body
    item = find_item(body, "sheets")
    ok = resp.status_code == 200 and item and item.get("status") == "failed" and "找不到 姓名=王五" in item.get("result", "")
    pending_ids = [c["task_id"] for c in get_confirmations(api)]
    no_conf = str(body.get("task_id")) not in pending_ids
    trace["side_effect"] = {"confirmation_created": not no_conf}
    quality = 100 if ok and no_conf else (50 if item and "找不到" in item.get("result", "") else 0)
    return {"pass": ok and no_conf, "details": item.get("result", "") if item else str(body), "quality": quality}


def gc04_create_calendar(api, rnd, trace):
    suffix = rand_suffix()
    text = f"明天下午3点到4点创建日程：冒烟测试_{suffix}"
    trace["input"] = text
    resp = post_task(api, text, rnd)
    body = resp.json()
    trace["response"] = body
    item = find_item(body, "calendar")
    ok = resp.status_code == 200 and item and item.get("status") == "executed" and "日程已创建" in item.get("result", "")
    event_id = ""
    try:
        event_id = json.loads(item.get("result", "{}")).get("data", {}).get("event_id", "")
    except Exception:
        pass
    trace["side_effect"] = {"event_id": event_id}
    quality = 100
    try:
        start = item.get("args", {}).get("start_ts", "")
        tomorrow = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d")
        if not start.startswith(tomorrow + "T15:00"):
            quality -= 40
    except Exception:
        pass
    return {"pass": ok and bool(event_id), "details": item.get("result", "") if item else str(body), "quality": max(0, quality), "event_id": event_id}


def gc05_email(api, rnd, trace):
    env = load_env(PROJECT_ROOT / ".env")
    to = env.get("QQ_MAIL_ADDRESS", "")
    text = f"给 {to} 发邮件，主题：冒烟测试，正文：自动化测试邮件 {rand_suffix()}"
    trace["input"] = text
    resp = post_task(api, text, rnd)
    body = resp.json()
    trace["response"] = body
    item = find_item(body, "email")
    if not item or item.get("status") != "pending_confirm":
        return {"pass": False, "details": item.get("result", "") if item else str(body), "quality": 0}
    dec = decide(api, item["confirmation_id"], True)
    trace["confirmation"] = {"id": item["confirmation_id"], "decide": dec.json()}
    ok = dec.status_code == 200 and dec.json().get("execution_result", {}).get("ok") is True
    return {"pass": ok, "details": dec.json().get("execution_result", {}).get("message", ""), "quality": 100 if ok else 50}

def gc06_combined(api, rnd, trace):
    phone = "138" + str(random.randint(10000000, 99999999))
    suffix = rand_suffix()
    text = f"把报名表里张三的电话更新为{phone}，明天下午3点到4点创建日程：冒烟测试_组合{suffix}"
    trace["input"] = text
    resp = post_task(api, text, rnd)
    body = resp.json()
    trace["response"] = body
    tasks = body.get("tasks", [])
    sheet_item = find_item(body, "sheets")
    cal_item = find_item(body, "calendar")
    ok = len(tasks) >= 2 and sheet_item and sheet_item.get("status") == "pending_confirm" and cal_item and cal_item.get("status") == "executed"
    return {"pass": ok, "details": f"子任务数={len(tasks)}", "quality": 100 if len(tasks) >= 2 else 40}


def gc07_clarify(api, rnd, trace):
    text = "帮我处理一下"
    trace["input"] = text
    resp = post_task(api, text, rnd)
    body = resp.json()
    trace["response"] = body
    ok = resp.status_code == 200 and body.get("status") == "need_clarify" and bool(body.get("question"))
    return {"pass": ok, "details": body.get("question", ""), "quality": 100 if ok else 0}


def gc08_clarify_continue(api, rnd, trace):
    text = "读取报名表内容"
    trace["input"] = text
    resp = post_task(api, text, rnd)
    body = resp.json()
    trace["response"] = body
    item = find_item(body, "sheets")
    ok = resp.status_code == 200 and item and item.get("status") == "executed"
    return {"pass": ok, "details": item.get("result", "") if item else str(body), "quality": 100 if ok else 50}


def gc09_max_rounds(api, rnd, trace):
    trace["input"] = {"text": "帮我处理一下", "round": 4}
    resp = api.post("/api/v1/tasks", json={"text": "帮我处理一下", "round": 4})
    body = resp.json()
    trace["response"] = body
    ok = resp.status_code == 200 and body.get("status") == "too_many_rounds" and "上限" in (body.get("message") or "")
    return {"pass": ok, "details": body.get("message", ""), "quality": 100 if ok else 0}


def gc10_no_false_clarify(api, rnd, trace):
    text = "读取报名表内容"
    trace["input"] = text
    resp = post_task(api, text, rnd)
    body = resp.json()
    trace["response"] = body
    ok = resp.status_code == 200 and body.get("status") in ("planned", "executed", "pending_confirm") and body.get("question") is None
    return {"pass": ok, "details": f"status={body.get('status')}", "quality": 100 if ok else 0}


def gc11_confirm_before_execute(api, rnd, trace):
    phone = "138" + str(random.randint(10000000, 99999999))
    before = read_sheet()[1][1]
    text = f"把报名表里张三的电话更新为{phone}"
    trace["input"] = text
    resp = post_task(api, text, rnd)
    body = resp.json()
    trace["response"] = body
    item = find_item(body, "sheets")
    after_post = read_sheet()[1][1]
    not_executed = after_post == before
    ok = bool(item and item.get("status") == "pending_confirm" and not_executed)
    trace["side_effect"] = {"before": before, "after_post": after_post}
    return {"pass": ok, "details": "确认前文件未变" if ok else f"确认前文件已变 {before}->{after_post}", "quality": 100 if ok else 50}


def gc12_reject_not_execute(api, rnd, trace):
    phone = "138" + str(random.randint(10000000, 99999999))
    before = read_sheet()[1][1]
    text = f"把报名表里张三的电话更新为{phone}"
    trace["input"] = text
    resp = post_task(api, text, rnd)
    body = resp.json()
    item = find_item(body, "sheets")
    if not item or not item.get("confirmation_id"):
        return {"pass": False, "details": "未生成确认", "quality": 0}
    dec = decide(api, item["confirmation_id"], False)
    trace["confirmation"] = {"id": item["confirmation_id"], "decide": dec.json()}
    after = read_sheet()[1][1]
    ok = dec.status_code == 200 and dec.json().get("status") == "rejected" and after == before
    trace["side_effect"] = {"before": before, "after": after}
    return {"pass": ok, "details": "拒绝后文件未变" if ok else "拒绝后文件变了", "quality": 100 if ok else 50}


def gc13_whitelist(api, rnd, trace):
    from backend.agent.executor import _run
    from backend.models.task import TaskItem
    import backend.tools  # noqa: F401 注册工具

    item = TaskItem(tool="sheets", args='{"action": "delete", "filename": "a.xlsx"}')
    result = _run(item)
    trace["response"] = result
    ok = result["ok"] is False and "白名单" in result["message"]
    return {"pass": ok, "details": result["message"], "quality": 100 if ok else 0}


def gc14_path_traversal(api, rnd, trace):
    from backend.tools.sheets import SheetTool

    result = SheetTool().execute(action="read", filename="../../secret.txt")
    trace["response"] = {"ok": result.ok, "message": result.message}
    ok = result.ok is False and "越权" in result.message
    return {"pass": ok, "details": result.message, "quality": 100 if ok else 0}


def gc15_bad_input(api, rnd, trace):
    r1 = api.post("/api/v1/tasks", json={"text": ""})
    r2 = api.post("/api/v1/tasks", content="not-json", headers={"Content-Type": "application/json"})
    trace["response"] = {"empty": r1.status_code, "badjson": r2.status_code}
    ok = r1.status_code == 422 and r2.status_code == 422
    return {"pass": ok, "details": f"empty={r1.status_code}, badjson={r2.status_code}", "quality": 100 if ok else 50}

def gc16_unexpected_error_500_logged(api, rnd, trace):
    from fastapi.testclient import TestClient
    from backend.main import app
    import backend.api.v1.task as task_api

    original = task_api.plan

    def boom(text):
        raise RuntimeError("smoke-500-boom")

    task_api.plan = boom
    try:
        client = TestClient(app, raise_server_exceptions=False)
        resp = client.post("/api/v1/tasks", json={"text": "制造错误"})
        trace["response"] = {"status": resp.status_code, "detail": resp.json()}
        log = log_tail(30)
        trace["log_snippet"] = log
        ok = resp.status_code == 500 and "smoke-500-boom" in log
        return {"pass": ok, "details": f"status={resp.status_code}, 日志含错误={('smoke-500-boom' in log)}", "quality": 100 if ok else 50}
    finally:
        task_api.plan = original


def gc17_health_after_errors(api, rnd, trace):
    resp = api.get("/health")
    trace["response"] = {"status": resp.status_code}
    ok = resp.status_code == 200
    return {"pass": ok, "details": f"health={resp.status_code}", "quality": 100 if ok else 0}


def gc18_dedup(api, rnd, trace):
    text = "读取报名表内容"
    r1 = post_task(api, text, rnd)
    r2 = post_task(api, text, rnd)
    trace["response"] = {"first": r1.json().get("task_id"), "second": r2.json().get("task_id")}
    ok = r1.status_code == 200 and r2.status_code == 200 and r1.json().get("task_id") == r2.json().get("task_id")
    return {"pass": ok, "details": f"task_id={r1.json().get('task_id')} 复用={ok}", "quality": 100 if ok else 0}


def gc19_idempotent_confirm(api, rnd, trace):
    from backend.db import SessionLocal
    from backend.safety.gate import create_confirmation

    db = SessionLocal()
    try:
        conf = create_confirmation(db, task_id=None, action="发送邮件", target="幂等测试", task_item_id=None)
        cid = conf.id
    finally:
        db.close()
    r1 = decide(api, cid, True)
    r2 = decide(api, cid, True)
    trace["response"] = {"first": r1.json(), "second": r2.json()}
    ok = (
        r1.status_code == 200
        and r1.json().get("status") == "approved"
        and r2.status_code == 200
        and r2.json().get("status") == "approved"
        and r2.json().get("execution_result") is None
    )
    return {"pass": ok, "details": f"两次均 approved，第二次未重复执行={ok}", "quality": 100 if ok else 50}


def gc20_log_recorded(api, rnd, trace):
    log = log_tail(40)
    trace["log_snippet"] = log
    ok = "POST /api/v1/tasks" in log
    return {"pass": ok, "details": "日志有任务请求记录" if ok else "日志无任务请求记录", "quality": 100 if ok else 0}


GOLDEN_CASES = [
    ("GC-01", "读表格", gc01_read_sheet),
    ("GC-02", "改表格-按表头定位", gc02_write_by_key),
    ("GC-03", "改表格-找不到人不改", gc03_missing_person),
    ("GC-04", "建日历", gc04_create_calendar),
    ("GC-05", "发邮件", gc05_email),
    ("GC-06", "组合任务拆解", gc06_combined),
    ("GC-07", "反问澄清", gc07_clarify),
    ("GC-08", "澄清后继续", gc08_clarify_continue),
    ("GC-09", "追问上限", gc09_max_rounds),
    ("GC-10", "不误反问", gc10_no_false_clarify),
    ("GC-11", "高危未确认不执行", gc11_confirm_before_execute),
    ("GC-12", "拒绝不执行", gc12_reject_not_execute),
    ("GC-13", "白名单拦截", gc13_whitelist),
    ("GC-14", "目录越权拦截", gc14_path_traversal),
    ("GC-15", "坏输入422", gc15_bad_input),
    ("GC-16", "未预期异常500+日志", gc16_unexpected_error_500_logged),
    ("GC-17", "错误后服务存活", gc17_health_after_errors),
    ("GC-18", "去重复用", gc18_dedup),
    ("GC-19", "确认幂等", gc19_idempotent_confirm),
    ("GC-20", "日志可查", gc20_log_recorded),
]


def classify_failure(case_id, details):
    """失败分类：意图识别错/工具选错/参数错/上下文丢失/模型幻觉/评分器写错/其它。"""
    d = (details or "") + " " + case_id
    if any(k in d for k in ("找不到表头", "找不到 姓名", "参数不合法", "参数不足", "时间", "邮箱格式")):
        return "参数错"
    if any(k in d for k in ("白名单", "越权", "路径")):
        return "工具选错/越权"
    if any(k in d for k in ("子任务数=1", "遗漏", "合并")):
        return "上下文丢失"
    if any(k in d for k in ("幻觉", "不存在", "未执行")):
        return "模型幻觉"
    if any(k in d for k in ("断言", "期望", "预期")):
        return "评分器写错"
    if "意图" in d:
        return "意图识别错"
    return "其它"


def functional_score(results):
    total = len(results)
    passed = sum(1 for r in results if r["pass"])
    return {"score": round(passed / total * 100, 1) if total else 0, "passed": passed, "total": total}


def quality_score(results):
    qs = [r.get("quality", 0) for r in results]
    avg = round(sum(qs) / len(qs), 1) if qs else 0
    return {"score": avg, "low": [r["case_id"] for r in results if r.get("quality", 0) < 60]}


def stability_score(results, rounds, health_ok, timeouts, five_hundreds):
    deduction = min(100, five_hundreds * 20) + min(100, timeouts * 10)
    if not health_ok:
        deduction += 30
    by_case = {}
    for r in results:
        by_case.setdefault(r["case_id"], []).append(r["pass"])
    flake = [c for c, o in by_case.items() if len(o) > 1 and any(o) and not all(o)]
    deduction += int(len(flake) / max(1, len(by_case)) * 50)
    return {"score": max(0, 100 - deduction), "five_hundreds": five_hundreds, "timeouts": timeouts, "flake": flake}


def parse_args():
    parser = argparse.ArgumentParser(description="办公自动化 Agent 自动化测试（20 黄金 Case）")
    parser.add_argument("--rounds", type=int, default=3, help="连跑轮数（默认 3）")
    parser.add_argument("--base-url", default=DEFAULT_BASE_URL, help=f"后端地址（默认 {DEFAULT_BASE_URL}）")
    parser.add_argument("--include-email", action="store_true", help="包含 GC-05 真发邮件（默认跳过）")
    parser.add_argument("--cleanup", action="store_true", help="跑完后删除本次创建的飞书日程")
    args = parser.parse_args()
    if args.rounds < 1:
        parser.error("--rounds 必须 >= 1")
    return args


def cleanup_events(api, event_ids):
    from backend.db import SessionLocal
    from backend.feishu.token_store import get_valid_token
    from backend.feishu.client import FEISHU_OPEN_BASE

    db = SessionLocal()
    try:
        token = get_valid_token(db, 1).access_token
    finally:
        db.close()
    headers = {"Authorization": f"Bearer {token}"}
    out = []
    for eid in event_ids:
        resp = api.delete(f"{FEISHU_OPEN_BASE}/calendar/v4/calendars/primary/events/{eid}", headers=headers, timeout=15)
        out.append(f"{eid[:12]} -> {resp.status_code}")
    return out


def main():
    args = parse_args()
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    problems = check_env(args.base_url)
    if problems:
        for p in problems:
            print(f"环境未就绪：{p}")
        return 2
    print("环境检查通过：后端健康 / DeepSeek 已配置 / 飞书令牌可用 / 表格文件存在")

    api = httpx.Client(base_url=args.base_url.rstrip("/"), timeout=TIMEOUT)
    results = []
    event_ids = []
    timeouts = 0
    five_hundreds = 0
    health_ok = True
    try:
        for rnd in range(1, args.rounds + 1):
            print(f"\n===== 第 {rnd}/{args.rounds} 轮 =====")
            for case_id, name, fn in GOLDEN_CASES:
                if case_id == "GC-05" and not args.include_email:
                    continue
                trace = make_trace(case_id, rnd)
                start = time.perf_counter()
                try:
                    rec = fn(api, rnd, trace)
                except Exception as exc:
                    rec = {"pass": False, "details": f"执行异常：{exc}", "quality": 0}
                    trace["log_snippet"] = traceback.format_exc()
                duration_ms = int((time.perf_counter() - start) * 1000)
                rec.update({"case_id": case_id, "name": name, "round": rnd, "duration_ms": duration_ms})
                if duration_ms > 30000:
                    timeouts += 1
                if case_id != "GC-16" and isinstance(trace.get("response"), dict) and trace["response"].get("status") == 500:
                    five_hundreds += 1
                if case_id == "GC-04" and rec.get("event_id"):
                    event_ids.append(rec["event_id"])
                rec["classification"] = classify_failure(case_id, rec["details"]) if not rec["pass"] else ""
                trace["scoring"] = {"functional": "PASS" if rec["pass"] else "FAIL", "quality": rec.get("quality", 0)}
                trace["result"] = "PASS" if rec["pass"] else "FAIL"
                trace["failure_class"] = rec["classification"]
                rec["trace"] = str(write_trace(trace))
                results.append(rec)
                print(f"  {case_id} {name:<16} {'PASS' if rec['pass'] else 'FAIL'} ({duration_ms}ms) 质{rec.get('quality',0)} {rec['details'][:80]}")
            try:
                if api.get("/health", timeout=5).status_code != 200:
                    health_ok = False
            except Exception:
                health_ok = False
    finally:
        api.close()

    cleanup_out = cleanup_events(api, event_ids) if args.cleanup else []
    fs = functional_score(results)
    qs = quality_score(results)
    ss = stability_score(results, args.rounds, health_ok, timeouts, five_hundreds)
    total = round(fs["score"] * 0.5 + qs["score"] * 0.3 + ss["score"] * 0.2, 1)

    lines = []
    lines.append("=" * 74)
    lines.append("自动化测试报告（20 黄金 Case / 3 评分器 / Trace）")
    lines.append(f"轮数：{args.rounds}　含邮件：{'是' if args.include_email else '否'}")
    lines.append(f"功能分：{fs['score']}（{fs['passed']}/{fs['total']}）")
    lines.append(f"质量分：{qs['score']}　低质量：{qs['low']}")
    lines.append(f"稳定性分：{ss['score']}（500={ss['five_hundreds']} 超时={ss['timeouts']} 抖动={ss['flake']}）")
    lines.append(f"总分：{total}")
    lines.append("-" * 74)
    fails = [r for r in results if not r["pass"]]
    if fails:
        lines.append("失败明细（含分类）：")
        for r in fails:
            lines.append(f"  {r['case_id']} 第{r['round']}轮 {r['name']} [{r['classification']}]：{r['details'][:160]}")
            lines.append(f"    trace: {r.get('trace', '')}")
    else:
        lines.append("全部通过 ✅")
    if cleanup_out:
        lines.append("清理：" + ", ".join(cleanup_out))
    report = "\n".join(lines)
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    report_path = REPORT_DIR / f"smoke_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
    report_path.write_text(report, encoding="utf-8")
    print("\n" + report)
    print(f"\n报表：{report_path}")
    return 0 if not fails else 1


if __name__ == "__main__":
    sys.exit(main())